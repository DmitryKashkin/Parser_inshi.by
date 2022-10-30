"""
1 получить список УРЛов
2 пройтись по нему в цикле

https://inshi.by/katalog/remmers

https://inshi.by/katalog/instrument
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
import json

exceptions_list = [
    'Кисти',
    'STORCH',
    'Инструмент для наливных полов',
    'Оборудование',
    'Крепеж Camo',
    'Строительство и ремонт',
    'Защита древесины',
    'Декоративные и промышленные наливные полы',
    'Защита и ремонт фасадов',
    'Интерьерные краски',
]

urls_list = [
    'https://inshi.by/katalog/remmers',
    'https://inshi.by/katalog/instrument/',
]

prefix = 'https://inshi.by/'

field_names = [
    'name',
    'description',
    'media',
    'technical_description',
]


def get_products_list(urls_list: list) -> list:
    products_list = []
    for item in urls_list:
        response = requests.get(item)
        soup = BeautifulSoup(response.text, 'lxml')
        products_list += soup.find_all('span', class_='prod-title')
    return products_list


def clear_products_list(products_list: list):
    cleared_list = []
    # for item in products_list:
    #     if itme in exceptions_list:
    #         continue
    #
    # print(products_list)
    for item in products_list:
        if item.text in exceptions_list:
            continue
        cleared_list.append(item)
    return cleared_list


def get_product(products_list, url_prefix=''):
    product = {}
    products = []
    for item in products_list:
        item = url_prefix + item.find('a').get('href')
        # print (item)
        response = requests.get(item)
        soup = BeautifulSoup(response.text, 'lxml')
        product['name'] = soup.find('h1').text
        product['description'] = ''
        descriptions = soup.find('div', class_="prod-desc js-product").find_all('p')
        for description in descriptions:
            product['description'] += description.text.replace('\r', '').replace('\n', '').replace('\t', '') + '\n'
        medias = soup.find_all('a', class_="fancybox")
        product['media'] = []
        for media in medias:
            product['media'].append(prefix + media.get('href'))
        technical_description = soup.find('div', class_="prod-desc js-product").find('a', target="_blank")
        if technical_description:
            product['technical_description'] = prefix + technical_description.get('href')
            # product['technical_description'] = soup.find('div', class_="prod-desc js-product").find('a', target="_blank").get('href')
        print(product)
        products.append(product)
        break
    return products


def save_to_excel(products):
    with open('products.json', 'w') as fp:
        json.dump(products, fp)
    # создаем новый excel-файл
    wb = openpyxl.Workbook()
    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    row = 2
    col = 1
    sheet.append(field_names)
    for item in products:
        for key, value in item.items():
            cell = sheet.cell(row=row, column=col)
            # print(value)
            if isinstance(value, list):
                value = '\n'.join(value)
            cell.value = value
            col += 1
        row += 1
    wb.save('product.xlsx')


def main():
    products_list = get_products_list(urls_list)
    products_list = clear_products_list(products_list)
    products = get_product(products_list, url_prefix=prefix)
    save_to_excel(products)


if __name__ == '__main__':
    main()
    # product = {}
    # item = 'https://inshi.by/katalog/remmers/zashhita-drevesiny/antiseptiki/aqua-ig-15-impagniergrund-it'
    # response = requests.get(item)
    # soup = BeautifulSoup(response.text, 'lxml')
    # product['name'] = soup.find('h1').text
    # product['description'] = ''
    # descriptions = soup.find('div', class_="prod-desc js-product").find_all('p')
    # for description in descriptions:
    #     product['description'] += description.text.replace('\r', '').replace('\n', '').replace('\t', '') + '\n'
    # print(product)
    # save_to_excel('test')
