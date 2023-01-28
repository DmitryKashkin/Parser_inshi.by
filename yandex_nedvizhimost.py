import asyncio
from time import sleep
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import openpyxl
from selenium.webdriver.common.by import By

URL = 'https://realty.ya.ru/moskva_i_moskovskaya_oblast/kupit/novostrojka/?showOutdated=NO'
PAGE = 'https://realty.ya.ru/moskva_i_moskovskaya_oblast/kupit/novostrojka/?showOutdated=NO&page='
CSS_CAPTCHA = '#root > div > div > form > div.Spacer.Spacer_auto-gap_bottom > div > div > div.CheckboxCaptcha-Anchor > input'
PREFIX = 'https://realty.ya.ru'
FILENAME = 'yandex_nedvizhimost.xlsx'


def main_captcha(driver, url, s):
    while True:
        driver.delete_all_cookies()
        driver.get(url)
        try:
            element = driver.find_element(By.CSS_SELECTOR, CSS_CAPTCHA)
        except:
            return
        driver.execute_script("arguments[0].click();", element)
        sleep(3)
        s.cookies.clear()
        [s.cookies.set(c['name'], c['value']) for c in driver.get_cookies()]
        response = s.get(url)
        if 'captcha' in response.text:
            continue
        break
    return response


def export_xls(spec_list):
    try:
        wb = openpyxl.load_workbook(FILENAME)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.create_sheet(title='Первый лист', index=0)
    ws = wb.worksheets[0]
    for row_data in spec_list:
        ws.append(list(row_data.values()))
        # print(list(row_data.values()))
    wb.save(FILENAME)


def get_spec(url, s, driver):
    spec = {'url': ' ',
            'realty_object': ' ',
            'deadline': ' ',
            'class': ' ',
            'home_type': ' ',
            'buildings': ' ',
            'finish': ' ',
            'queues': ' ',
            'num_of_apart': ' ',
            }
    while True:
        response = s.get(url)
        if response:
            break
        continue
    if 'captcha' in response.text:
        response = main_captcha(driver, url, s)
    soup = BeautifulSoup(response.text, 'lxml')
    spec['url'] = url
    spec['realty_object'] = soup.find('h1').text.replace('\xa0', ' ')
    description = soup.find('h2', class_='SiteCardDescription__title--35WGe')
    about_object = soup.find('h2', string='Об объекте')
    if about_object:
        try:
            spec['deadline'] = about_object.next_sibling.find('span', string='Срок сдачи').next_sibling.text.replace(
                '\xa0', ' ')
        except:
            ...
        try:
            spec['class'] = about_object.next_sibling.find('span', string='Класс жилья').next_sibling.text.replace(
                '\xa0', ' ')
        except:
            ...
    try:
        spec['home_type'] = description.next_sibling.find('div', string='Тип дома').next_sibling.text.replace('\xa0',
                                                                                                              ' ')
    except:
        ...
    try:
        spec['buildings'] = description.next_sibling.find('div', string='Число корпусов').next_sibling.text.replace(
            '\xa0', ' ')
    except:
        ...
    try:
        spec['finish'] = description.next_sibling.find('div', string='Отделка').next_sibling.text.replace('\xa0', ' ')
    except:
        ...
    try:
        spec['queues'] = description.next_sibling.find('div', string='Очереди').next_sibling.text.replace('\xa0', ' ')
    except:
        ...
    try:
        spec['num_of_apart'] = description.next_sibling.find('div', string='Число квартир').next_sibling.text.replace(
            '\xa0', ' ')
    except:
        try:
            spec['num_of_apart'] = description.next_sibling.find('div',
                                                                 string='Число квартир и апартаментов').next_sibling.text.replace(
                '\xa0', ' ')
        except:
            try:
                spec['num_of_apart'] = description.next_sibling.find('div',
                                                                     string='Число апартаментов').next_sibling.text.replace(
                    '\xa0', ' ')
            except:
                ...

    print(spec)
    return spec


def main():
    url_list = []
    try:
        wb = openpyxl.load_workbook(FILENAME)
        ws = wb.worksheets[0]
        for row in range(1, ws.max_row + 1):
            url_list.append(ws.cell(row, 1).value)
    except FileNotFoundError:
        ...
    s = requests.Session()
    driver = webdriver.Chrome()
    response = main_captcha(driver, URL, s)
    soup = BeautifulSoup(response.text, 'lxml')
    page_number = 1
    while True:
        spec_list = []
        print('!!!!!page_number ', page_number)
        for item in soup.find_all('div', class_='SiteSnippetSearch SitesSerp__snippet'):
            url = PREFIX + item.a.get('href')
            print(url)
            if url in url_list:
                print('skip')
                continue
            spec_list.append(get_spec(url, s, driver))
        export_xls(spec_list)
        page_number += 1
        url = PAGE + str(page_number)
        print(url)
        response = s.get(url)
        if 'captcha' in response.text:
            response = main_captcha(driver, url, s)
        if response.status_code == 404:
            break
        soup = BeautifulSoup(response.text, 'lxml')


if __name__ == '__main__':
    main()
