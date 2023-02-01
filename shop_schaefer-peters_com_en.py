from time import sleep
import openpyxl
import requests
from bs4 import BeautifulSoup
import fake_useragent
import re

SUFFIX = '?pgNr='
URL = 'https://shop.schaefer-peters.com/en/'
user = fake_useragent.UserAgent().random
header = {
    'user-agent': user
}
pattern = re.compile(r'tab\d')
FILENAME = 'shop.schaefer-peters.xlsx'


def export_xls(item_list):
    header = {}
    try:
        wb = openpyxl.load_workbook(FILENAME)
        ws = wb.worksheets[0]
        for col in range(1, ws.max_column + 1): # Получаем заголовок таблицы
            # header.append(ws.cell(1,col).value)
            header[ws.cell(1, col).value] = col
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.create_sheet(title='Первый лист', index=0)
    ws = wb.worksheets[0]
    max_row = ws.max_row
    for row_data in item_list:
        for key, value in row_data.items():
            try:
                header[key]
            except KeyError:
                try:
                    header[key] = max(header.values()) + 1
                except ValueError:
                    header[key] = 1
            ws.cell(max_row + 1, header[key]).value = value
        # ws.append(list(row_data.values()))
        # print(list(row_data.values()))
        max_row += 1
    for key, value in header.items(): # обновляем заголовок таблицы
        ws.cell(1,value).value = key
    wb.save(FILENAME)


def get_item(cat_list, s):
    exclusion_list = []
    try:  # check downloaded items
        wb = openpyxl.load_workbook(FILENAME)
        ws = wb.worksheets[0]
        for row in range(1, ws.max_row + 1):
            exclusion_list.append(ws.cell(row, 4).value)
    except FileNotFoundError:
        ...

    for url in cat_list:
        suffix = ''
        page = 0
        while True:  # получаем данные постранично
            item_list = []
            response = s.get(url + suffix)
            if response.url != url + suffix:
                break
            soup = BeautifulSoup(response.text, 'lxml')
            tbody = soup.find('div', class_='list-container')
            for item in tbody.find_all('tr', class_='sp-product-item'):
                item_url = item.find('a').get('href')
                print(item_url)
                if item_url in exclusion_list:
                    print('skip')
                    continue
                item_list.append(get_spec(item_url, s))
            export_xls(item_list)
            page += 1
            suffix = SUFFIX + str(page)
            # sleep(1)


def get_spec(url, s):
    item_spec = {
        'art': ' ',
        'art2': ' ',
        'cat_path': ' ',
        'url': ' ',
        'img_urls': ' ',
    }
    keys = {}
    response = s.get(url)
    soup = BeautifulSoup(response.text, 'lxml')
    item_spec['cat_path'] = soup.find('div', id='breadcrumb').text
    item_spec['url'] = url
    content_box = soup.find('div', class_='content-box')
    art = content_box.find('div', class_='col-xs-12 attribute-title')
    item_spec['art'] = art.h1.text
    item_spec['art2'] = art.span.text
    img_urls = content_box.find('div', class_='details-image-gallery')
    item_spec['img_urls'] = ', '.join(
        [i.get('data-src') for i in img_urls.find_all('div', class_='item') if i.get('data-src')])
    custom_attributes = content_box.find('div', class_='custom-attributes').find_all('div', class_='row')
    for attrib in custom_attributes:
        if attrib.div.span.strong:
            key = attrib.contents[1].text.replace('\n', '')
            value = attrib.contents[3].text.replace('\n', '').replace(' ', '')
            item_spec[key] = value
    # relatedInfo_relatedInfoFull = content_box.find('div', class_='tab-content')
    relatedInfo_relatedInfoFull = content_box.find('div', class_='relatedInfo relatedInfoFull')
    # for attrib in relatedInfo_relatedInfoFull.find_all('div', class_='row'):
    # attribs = relatedInfo_relatedInfoFull.find_all('div', {'id': re.compile(r'tab\d')})
    for attrib in relatedInfo_relatedInfoFull.find('div', id='tab1').find_all('div', class_='row'):
        key = attrib.contents[1].text.replace('\n', '')
        value = attrib.contents[3].text.replace('\n', '').replace(' ', '')
        item_spec[key] = value
    for attrib in relatedInfo_relatedInfoFull.find('ul', class_="responsive-tabs nav nav-tabs").find_all('li'):
        key = attrib.a.text.replace('\n', '')
        key0 = attrib.a.get('href')[1:]
        keys[key0] = key
        value = ''
        item_spec[key] = value

    for attrib in relatedInfo_relatedInfoFull.find_all('div', {'id': re.compile(r'tab\d')}):
        if attrib.get('id') == 'tab1':
            continue
        key0 = attrib.get('id')
        key = keys[key0]
        value = attrib.text.replace('\n', '')
        item_spec[key] = value

    # item_spec['application_area'] = relatedInfo_relatedInfoFull.find('div', class_="tab-pane", id='tab4').text.replace(
    #     '\n', '')
    # catalogue_pages = relatedInfo_relatedInfoFull.find('a', class_='sx-product-download').get('data-filename')
    # item_spec['catalogue_pages'] = url + catalogue_pages
    # # url0 = 'https://shop.schaefer-peters.com/DIN-80704-A4-M-10/'
    # url0 = 'https://shop.schaefer-peters.com/index.php?lang=1&amp'
    # form_data = {'stoken': "CA4B2959",
    #              'lang': "1",
    #              'fnc': "sxdownloadpdffile",
    #              'cl': "details",
    #              'guid': "a22f0b16-97f4-410e-982d-46b2f6947f6c",
    #              'filename': "Gesamtkatalog DE-EN 598.pdf",
    #              'anid': "56a7d0df1a3c39d35217ce48f9d246ca"
    #              }
    # pdf = requests.post(url0, data=form_data, headers=header)
    # # pdf = s.post(url0, data=form_data)
    #
    # # pdf = s.get(item_spec['catalogue_pages'])
    # sleep(3)
    return item_spec


def get_cat(url, s):
    cat_list = []
    response = s.get(url)
    soup = BeautifulSoup(response.text, 'lxml')
    for cat_menu in soup.find_all('ul', class_='level-4'):
        cat_list.append(cat_menu.find('a').get('href'))
        # print(cat_menu.find('a').get('href'))
    return cat_list


def main():
    s = requests.Session()
    cat_list = get_cat(URL, s)
    get_item(cat_list, s)


if __name__ == '__main__':
    main()
